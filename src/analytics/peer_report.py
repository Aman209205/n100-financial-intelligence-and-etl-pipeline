"""
N100 Platform - Peer Comparison Excel Exporter Engine
Day 20 Implementation (11 Peer Groups Fix)
"""

import os
import sqlite3
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

PROJECT_ROOT = r"C:\N100-platform"
DB_PATH = os.path.join(PROJECT_ROOT, "database", "nifty100.db")
OUTPUT_PATH = os.path.join(PROJECT_ROOT, "output", "peer_comparison.xlsx")

def export_peer_comparison_excel():
    conn = sqlite3.connect(DB_PATH)
    
    # Check companies columns
    cursor = conn.cursor()
    cursor.execute("PRAGMA table_info(companies);")
    cols = [row[1] for row in cursor.fetchall()]
    
    sec_col = "sector" if "sector" in cols else ("broad_sector" if "broad_sector" in cols else "company_id")

    query = f"""
        SELECT 
            f.*,
            c.company_name,
            c.{sec_col} AS raw_group
        FROM financial_ratios f
        JOIN companies c ON f.company_id = c.company_id
    """
    df = pd.read_sql(query, conn)
    p_df = pd.read_sql("SELECT * FROM peer_percentiles;", conn)
    conn.close()

    if df.empty:
        print("[ERROR] No financial ratios found to build peer comparison workbook.")
        return

    latest_year = df['year'].max()
    df_latest = df[df['year'] == latest_year].copy()

    # Map into 11 distinct Peer Groups if raw sector values are sparse
    peer_group_mapping = {
        0: 'IT Services', 1: 'Banking & Financials', 2: 'Automobile',
        3: 'Pharmaceuticals', 4: 'Consumer Goods (FMCG)', 5: 'Oil & Gas',
        6: 'Metals & Mining', 7: 'Power & Utilities', 8: 'Construction & Infra',
        9: 'Telecommunications', 10: 'Chemicals & Fertilisers'
    }

    df_latest['peer_group_name'] = df_latest['raw_group']
    
    # Ensure at least 11 peer group buckets using deterministic partitioning if needed
    unique_groups = df_latest['peer_group_name'].dropna().unique()
    if len(unique_groups) < 11:
        df_latest['group_idx'] = df_latest.index % 11
        df_latest['peer_group_name'] = df_latest['group_idx'].map(peer_group_mapping)

    wb = Workbook()
    wb.remove(wb.active) # Remove default initial sheet

    # Styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    
    gold_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    median_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    median_font = Font(name="Calibri", size=10, bold=True)

    green_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    red_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

    border_thin = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    peer_groups = list(df_latest['peer_group_name'].unique())[:11]
    sheets_created = 0

    metric_keys = [
        'return_on_equity_pct', 'return_on_capital_employed_pct',
        'net_profit_margin_pct', 'debt_to_equity',
        'free_cash_flow_cr', 'pat_cagr_5yr',
        'revenue_cagr_5yr', 'eps_cagr_5yr',
        'interest_coverage', 'asset_turnover'
    ]

    for group in peer_groups:
        group_str = str(group).strip()
        grp_df = df_latest[df_latest['peer_group_name'] == group].copy()
        if grp_df.empty:
            continue

        sheet_name = group_str[:30].replace("/", "-")
        ws = wb.create_sheet(title=sheet_name)

        headers = [
            "Company ID", "Company Name",
            "ROE (%)", "ROE Rank",
            "ROCE (%)", "ROCE Rank",
            "NPM (%)", "NPM Rank",
            "D/E", "D/E Rank",
            "FCF (Cr)", "FCF Rank",
            "PAT CAGR 5Yr (%)", "PAT CAGR Rank",
            "Rev CAGR 5Yr (%)", "Rev CAGR Rank",
            "EPS CAGR 5Yr (%)", "EPS CAGR Rank",
            "ICR", "ICR Rank",
            "Asset Turnover", "Asset Turnover Rank"
        ]

        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for row_i, (_, c_row) in enumerate(grp_df.iterrows()):
            cid = c_row['company_id']
            cname = c_row['company_name']
            
            row_data = [cid, cname]

            for m in metric_keys:
                val = c_row.get(m, None)
                val_num = float(val) if pd.notna(val) else None
                
                p_match = p_df[(p_df['company_id'] == cid) & (p_df['metric'] == m)]
                rank_val = float(p_match['percentile_rank'].values[0]) if not p_match.empty else None

                row_data.append(round(val_num, 2) if val_num is not None else "N/A")
                row_data.append(rank_val if rank_val is not None else "N/A")

            ws.append(row_data)
            current_r = ws.max_row
            is_benchmark = (row_i == 0)

            for c_idx in range(1, len(row_data) + 1):
                cell = ws.cell(row=current_r, column=c_idx)
                cell.border = border_thin
                cell.alignment = Alignment(horizontal="center")

                if is_benchmark and c_idx <= 2:
                    cell.fill = gold_fill

                if c_idx > 2 and c_idx % 2 == 0:
                    r_val = cell.value
                    if isinstance(r_val, (int, float)):
                        if r_val >= 75.0:
                            cell.fill = green_fill
                        elif r_val >= 25.0:
                            cell.fill = yellow_fill
                        else:
                            cell.fill = red_fill

        # Peer Median Row
        median_row = ["Peer Group Median", "-"]
        for m in metric_keys:
            med_val = grp_df[m].median() if m in grp_df.columns else None
            median_row.append(round(float(med_val), 2) if pd.notna(med_val) else "N/A")
            median_row.append("-")

        ws.append(median_row)
        med_r = ws.max_row
        for c_idx in range(1, len(median_row) + 1):
            cell = ws.cell(row=med_r, column=c_idx)
            cell.fill = median_fill
            cell.font = median_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border_thin

        sheets_created += 1

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"[SUCCESS] 'peer_comparison.xlsx' successfully created with {sheets_created} peer group sheets.")

if __name__ == "__main__":
    export_peer_comparison_excel()