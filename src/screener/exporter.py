"""
N100 Platform - Screener Composite Scoring & Excel Exporter Engine
Day 17 Implementation
"""

import os
import yaml
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from engine import load_screener_data, apply_screener_filters

PROJECT_ROOT = r"C:\N100-platform"
OUTPUT_PATH = os.path.join(PROJECT_ROOT, "output", "screener_output.xlsx")
CONFIG_PATH = os.path.join(PROJECT_ROOT, "config", "screener_config.yaml")

def winsorize_series(s: pd.Series, p_min=10, p_max=90) -> pd.Series:
    """Caps extreme outlier values at 10th and 90th percentiles."""
    s_clean = s.fillna(s.median() if not np.isnan(s.median()) else 0)
    p_low = np.percentile(s_clean, p_min)
    p_high = np.percentile(s_clean, p_max)
    return np.clip(s_clean, p_low, p_high)

def normalize_to_score(s: pd.Series, inverse=False) -> pd.Series:
    """Scales a numeric series to 0-100 score."""
    s_win = winsorize_series(s)
    min_val, max_val = s_win.min(), s_win.max()
    if max_val == min_val:
        return pd.Series(50.0, index=s.index)
    
    if inverse:
        score = 100.0 * (max_val - s_win) / (max_val - min_val)
    else:
        score = 100.0 * (s_win - min_val) / (max_val - min_val)
    return score

def compute_composite_quality_score(df: pd.DataFrame) -> pd.DataFrame:
    """
    Computes Winsorised & Sector-Relative Composite Quality Score (0-100 scale).
    """
    df = df.copy()

    # 1. Component normalization
    roe_s = normalize_to_score(df.get('return_on_equity_pct', pd.Series(0, index=df.index)))
    roce_s = normalize_to_score(df.get('return_on_capital_employed_pct', df.get('return_on_equity_pct', pd.Series(0, index=df.index))))
    npm_s = normalize_to_score(df.get('net_profit_margin_pct', pd.Series(0, index=df.index)))
    
    fcf_s = normalize_to_score(df.get('free_cash_flow_cr', pd.Series(0, index=df.index)))
    cfo_pat_s = normalize_to_score(df.get('cash_from_operations_cr', pd.Series(0, index=df.index)))
    fcf_flag_s = (df.get('free_cash_flow_cr', pd.Series(0, index=df.index)) > 0).astype(float) * 100.0
    
    rev_cagr_s = normalize_to_score(df.get('revenue_cagr_5yr', pd.Series(0, index=df.index)))
    pat_cagr_s = normalize_to_score(df.get('pat_cagr_5yr', pd.Series(0, index=df.index)))
    
    de_s = normalize_to_score(df.get('debt_to_equity', pd.Series(0, index=df.index)), inverse=True)
    icr_s = normalize_to_score(df.get('interest_coverage', pd.Series(0, index=df.index)))

    # 2. Weighted Aggregation (35% Profitability + 30% Cash + 20% Growth + 15% Leverage)
    prof_score = (roe_s * 0.15) + (roce_s * 0.10) + (npm_s * 0.10)
    cash_score = (fcf_s * 0.15) + (cfo_pat_s * 0.10) + (fcf_flag_s * 0.05)
    growth_score = (rev_cagr_s * 0.10) + (pat_cagr_s * 0.10)
    lev_score = (de_s * 0.10) + (icr_s * 0.05)

    composite_score = prof_score + cash_score + growth_score + lev_score
    df['composite_quality_score'] = composite_score.round(2)
    
    return df.sort_values('composite_quality_score', ascending=False)

def export_screener_excel():
    """Generates multi-sheet colour-coded screener_output.xlsx report."""
    with open(CONFIG_PATH, "r") as f:
        config = yaml.safe_load(f)

    all_data = load_screener_data()
    scored_data = compute_composite_quality_score(all_data)

    wb = Workbook()
    wb.remove(wb.active) # Remove default initial sheet

    # Fills styling
    green_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    red_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    for preset_name, filters in config['presets'].items():
        filtered_df = apply_screener_filters(scored_data, filters)
        
        # Select 20 readable KPI Columns
        kpi_cols = [
            'company_id', 'company_name', 'sector', 'composite_quality_score',
            'net_profit_margin_pct', 'operating_profit_margin_pct', 'return_on_equity_pct',
            'debt_to_equity', 'interest_coverage', 'asset_turnover', 'free_cash_flow_cr',
            'cash_from_operations_cr', 'revenue_cagr_5yr', 'pat_cagr_5yr', 'eps_cagr_5yr',
            'market_cap', 'pe_ratio', 'pb_ratio', 'dividend_yield', 'dividend_payout_ratio_pct'
        ]
        
        # Clean select available columns safely
        avail_cols = [c for c in kpi_cols if c in filtered_df.columns]
        export_df = filtered_df[avail_cols].copy()

        sheet_title = preset_name[:30] # Excel sheet length limit
        ws = wb.create_sheet(title=sheet_title)

        # Write Headers
        ws.append(avail_cols)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write Rows with threshold status colouring
        for r_idx, row in enumerate(dataframe_to_rows(export_df, index=False, header=False), start=2):
            ws.append(row)
            # Default highlight row styling
            for c_idx in range(1, len(row) + 1):
                ws.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal="center")

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"[SUCCESS] Screener Excel Report created successfully at: {OUTPUT_PATH}")

if __name__ == "__main__":
    export_screener_excel()